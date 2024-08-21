import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

# Create a workbook and select the active worksheet
wb = Workbook()
ws = wb.active

# Set the header row
headers = ["Total", "Rating", "Time", "DateDebut", "DateFin", "MissionType", "Saison", "TypeDrone"]
ws.append(headers)

# Sample data
data = [
    [100, 4.5, datetime.now().strftime("%H:%M:%S"), "2024-07-01", "2024-07-02", "Survey", "Ete", "DroneA"],
    [200, 3.8, datetime.now().strftime("%H:%M:%S"), "2024-07-03", "2024-07-04", "Mapping", "Hiver", "DroneB"],
    [150, 4.2, datetime.now().strftime("%H:%M:%S"), "2024-07-05", "2024-07-06", "Inspection", "Printemps", "DroneC"],
]

# Append sample data to the worksheet
for row in data:
    ws.append(row)

# Adjust column widths
for col in range(1, len(headers) + 1):
    ws.column_dimensions[get_column_letter(col)].width = 15

# Save the workbook
file_path = "missions.xlsx"
wb.save(file_path)
print(f"File saved as {file_path}")
